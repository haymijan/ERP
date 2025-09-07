# stock/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from .models import Warehouse, Location
from .forms import WarehouseForm, LocationForm
from .models import LotSerialNumber
from .forms import LotSerialNumberForm
from .models import InventoryTransaction, Stock
from .forms import InventoryTransactionForm, InventoryAdjustmentForm
from django.db import transaction
from django.core.exceptions import ValidationError
from django.http import JsonResponse
from .forms import DateRangeForm
from datetime import timedelta
from django.http import JsonResponse
from django.db.models import Q
from django.core.paginator import Paginator
from .forms import StockMovementFilterForm
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from .forms import TransactionFilterForm
from django.db.models import Sum
from products.models import Product


# --- Warehouse CRUD Views ---
@login_required
@permission_required('stock.view_warehouse', login_url='/admin/')
def warehouse_list(request):
    warehouses = Warehouse.objects.all().order_by('name')
    context = {'warehouses': warehouses, 'title': 'All Warehouses'}
    return render(request, 'stock/warehouse_list.html', context)

@login_required
@permission_required('stock.add_warehouse', login_url='/admin/')
def add_warehouse(request):
    if request.method == 'POST':
        form = WarehouseForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('stock:warehouse_list')
    else:
        form = WarehouseForm()
    return render(request, 'stock/add_warehouse.html', {'form': form, 'title': 'Add New Warehouse'})

@login_required
@permission_required('stock.change_warehouse', login_url='/admin/')
def edit_warehouse(request, pk):
    warehouse = get_object_or_404(Warehouse, pk=pk)
    if request.method == 'POST':
        form = WarehouseForm(request.POST, instance=warehouse)
        if form.is_valid():
            form.save()
            return redirect('stock:warehouse_list')
    else:
        form = WarehouseForm(instance=warehouse)
    return render(request, 'stock/edit_warehouse.html', {'form': form, 'title': f'Edit Warehouse: {warehouse.name}'})

@login_required
@permission_required('stock.delete_warehouse', login_url='/admin/')
def delete_warehouse(request, pk):
    warehouse = get_object_or_404(Warehouse, pk=pk)
    if request.method == 'POST':
        warehouse.delete()
        return redirect('stock:warehouse_list')
    return render(request, 'confirm_delete.html', {'object': warehouse, 'title': f'Confirm Delete Warehouse: {warehouse.name}'})

# --- Location CRUD Views ---
@login_required
@permission_required('stock.view_location', login_url='/admin/')
def location_list(request):
    locations = Location.objects.all().order_by('name')
    context = {'locations': locations, 'title': 'All Locations'}
    return render(request, 'stock/location_list.html', context)

@login_required
@permission_required('stock.add_location', login_url='/admin/')
def add_location(request):
    if request.method == 'POST':
        form = LocationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('stock:location_list')
    else:
        form = LocationForm()
    return render(request, 'stock/add_location.html', {'form': form, 'title': 'Add New Location'})

@login_required
@permission_required('stock.change_location', login_url='/admin/')
def edit_location(request, pk):
    location = get_object_or_404(Location, pk=pk)
    if request.method == 'POST':
        form = LocationForm(request.POST, instance=location)
        if form.is_valid():
            form.save()
            return redirect('stock:location_list')
    else:
        form = LocationForm(instance=location)
    return render(request, 'stock/edit_location.html', {'form': form, 'title': f'Edit Location: {location.name}'})

@login_required
@permission_required('stock.delete_location', login_url='/admin/')
def delete_location(request, pk):
    location = get_object_or_404(Location, pk=pk)
    if request.method == 'POST':
        location.delete()
        return redirect('stock:location_list')
    return render(request, 'confirm_delete.html', {'object': location, 'title': f'Confirm Delete Location: {location.name}'})

@login_required
@permission_required('stock.view_lotserialnumber', login_url='/admin/')
def lot_serial_list(request):
    lots = LotSerialNumber.objects.select_related('product', 'location').all()
    context = {'lots': lots, 'title': 'All Lot/Serial Numbers'}
    return render(request, 'stock/lot_serial_list.html', context)

@login_required
@permission_required('stock.add_lotserialnumber', login_url='/admin/')
def add_lot_serial(request):
    if request.method == 'POST':
        form = LotSerialNumberForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('stock:lot_serial_list')
    else:
        form = LotSerialNumberForm()
    context = {'form': form, 'title': 'Add New Lot/Serial Number'}
    return render(request, 'stock/add_lot_serial.html', context)

@login_required
@permission_required('stock.change_lotserialnumber', login_url='/admin/')
def edit_lot_serial(request, pk):
    lot = get_object_or_404(LotSerialNumber, pk=pk)
    if request.method == 'POST':
        form = LotSerialNumberForm(request.POST, instance=lot)
        if form.is_valid():
            form.save()
            return redirect('stock:lot_serial_list')
    else:
        form = LotSerialNumberForm(instance=lot)
    context = {'form': form, 'title': f'Edit Lot/Serial: {lot.lot_number}'}
    return render(request, 'stock/edit_lot_serial.html', context)

@login_required
@permission_required('stock.delete_lotserialnumber', login_url='/admin/')
def delete_lot_serial(request, pk):
    lot = get_object_or_404(LotSerialNumber, pk=pk)
    if request.method == 'POST':
        lot.delete()
        return redirect('stock:lot_serial_list')
    return render(request, 'confirm_delete.html', {'object': lot, 'title': f'Confirm Delete Lot/Serial: {lot.lot_number}'})

@login_required
def transaction_list(request):
    transactions_queryset = InventoryTransaction.objects.select_related(
        'product', 'warehouse', 'source_location__warehouse', 'destination_location__warehouse', 'lot_serial', 'user'
    ).order_by('-transaction_date')

    user = request.user
    if not user.is_superuser:
        user_warehouse = getattr(user, 'warehouse', None)
        if user_warehouse:
            # ================== পরিবর্তিত অংশ শুরু ==================
            # এই কুয়েরিটি এখন সরাসরি warehouse ফিল্ডও পরীক্ষা করবে
            transactions_queryset = transactions_queryset.filter(
                Q(source_location__warehouse=user_warehouse) | 
                Q(destination_location__warehouse=user_warehouse) |
                Q(warehouse=user_warehouse)
            ).distinct()
            # ================== পরিবর্তিত অংশ শেষ ===================

    form = TransactionFilterForm(request.GET, user=request.user)
    if form.is_valid():
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        selected_user = form.cleaned_data.get('user')
        warehouse = form.cleaned_data.get('warehouse')

        if start_date:
            transactions_queryset = transactions_queryset.filter(transaction_date__date__gte=start_date)
        if end_date:
            transactions_queryset = transactions_queryset.filter(transaction_date__date__lte=end_date)
        if selected_user:
            transactions_queryset = transactions_queryset.filter(user=selected_user)
        if warehouse and user.is_superuser:
            # অ্যাডমিনের ফিল্টারেও warehouse ফিল্ডটি যোগ করা হলো
            transactions_queryset = transactions_queryset.filter(
                Q(source_location__warehouse=warehouse) | 
                Q(destination_location__warehouse=warehouse) |
                Q(warehouse=warehouse)
            ).distinct()

    paginator = Paginator(transactions_queryset, 20)
    page_number = request.GET.get('page')
    transactions = paginator.get_page(page_number)

    context = {
        'title': 'All Transactions',
        'transactions': transactions,
        'form': form,
    }
    return render(request, 'stock/transaction_list.html', context)

@login_required
@permission_required('stock.add_inventorytransaction', login_url='/admin/')
def record_transaction(request):
    if request.method == 'POST':
        form = InventoryTransactionForm(request.POST)
        if form.is_valid():
            # This is a simplified version. The full logic for stock updates needs to be added here.
            form.save()
            return redirect('stock:transaction_list')
    else:
        form = InventoryTransactionForm()
    return render(request, 'stock/record_transaction.html', {'form': form, 'title': 'Record Inventory Transaction'})

@login_required
@permission_required('stock.add_inventorytransaction', login_url='/admin/')
def inventory_adjustment(request):
    if request.method == 'POST':
        form = InventoryAdjustmentForm(request.POST)
        if form.is_valid():
            product = form.cleaned_data['product']
            location = form.cleaned_data['location']
            new_quantity = form.cleaned_data['new_quantity']
            with transaction.atomic():
                stock, created = Stock.objects.get_or_create(product=product, location=location, defaults={'quantity': 0})
                current_quantity = stock.quantity
                adjustment_quantity = new_quantity - current_quantity
                if adjustment_quantity != 0:
                    InventoryTransaction.objects.create(
                        product=product, 
                        transaction_type='adjustment', 
                        quantity=adjustment_quantity, 
                        destination_location=location if adjustment_quantity > 0 else None, 
                        source_location=location if adjustment_quantity < 0 else None, 
                        notes=f"Inventory adjustment. Old: {current_quantity}, New: {new_quantity}"
                    )
                    stock.quantity = new_quantity
                    stock.save()
            return redirect('stock:transaction_list')
    else:
        form = InventoryAdjustmentForm()
    context = {'form': form, 'title': 'Inventory Adjustment'}
    return render(request, 'stock/inventory_adjustment.html', context)

@login_required
@permission_required('stock.change_inventorytransaction', login_url='/admin/')
def edit_transaction(request, pk):
    transaction_obj = get_object_or_404(InventoryTransaction, pk=pk)
    if request.method == 'POST':
        form = InventoryTransactionForm(request.POST, instance=transaction_obj)
        if form.is_valid():
            form.save()
            return redirect('stock:transaction_list')
    else:
        form = InventoryTransactionForm(instance=transaction_obj)
    return render(request, 'stock/edit_transaction.html', {'form': form, 'title': f'Edit Transaction: {transaction_obj.product.name}'})

@login_required
@permission_required('stock.delete_inventorytransaction', login_url='/admin/')
def delete_transaction(request, pk):
    transaction_obj = get_object_or_404(InventoryTransaction, pk=pk)
    if request.method == 'POST':
        transaction_obj.delete()
        return redirect('stock:transaction_list')
    return render(request, 'confirm_delete.html', {'object': transaction_obj, 'title': 'Confirm Delete Transaction'})

@login_required
def check_product_tracking(request, product_id):
    try:
        product = Product.objects.get(pk=product_id)
        data = {'tracking_method': product.tracking_method}
        if product.tracking_method in ['lot', 'serial']:
            lots = LotSerialNumber.objects.filter(product=product, quantity__gt=0).values('id', 'lot_number', 'quantity')
            data['lots'] = list(lots)
        return JsonResponse(data)
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)
    
@login_required
def stock_movement_report(request):
    """
    এই ভিউটি এখন ব্যবহারকারীর ব্রাঞ্চ অনুযায়ী স্টক মুভমেন্ট দেখাবে এবং পেজিনেশন যোগ করবে।
    """
    transactions_list = InventoryTransaction.objects.select_related(
        'product', 'source_location__warehouse', 'destination_location__warehouse'
    ).order_by('-transaction_date')

    # --- ব্যবহারকারীর ব্রাঞ্চ অনুযায়ী ফিল্টারিং ---
    user = request.user
    if not user.is_superuser:
        user_warehouse = getattr(user, 'warehouse', None)
        if user_warehouse:
            # শুধুমাত্র সেইসব লেনদেন দেখানো হবে যার উৎস বা গন্তব্য ব্যবহারকারীর ব্রাঞ্চ
            transactions_list = transactions_list.filter(
                Q(source_location__warehouse=user_warehouse) |
                Q(destination_location__warehouse=user_warehouse)
            )

    # --- ফর্ম ফিল্টার (আপনার form.py অনুযায়ী) ---
    form = StockMovementFilterForm(request.GET)
    if form.is_valid():
        product = form.cleaned_data.get('product')
        transaction_type = form.cleaned_data.get('transaction_type')
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        warehouse = form.cleaned_data.get('warehouse')

        if product:
            transactions_list = transactions_list.filter(product=product)
        if transaction_type:
            transactions_list = transactions_list.filter(transaction_type=transaction_type)
        if start_date:
            transactions_list = transactions_list.filter(transaction_date__date__gte=start_date)
        if end_date:
            transactions_list = transactions_list.filter(transaction_date__date__lte=end_date)
        
        # সুপারইউজার ফিল্টার করলে সব ব্রাঞ্চ দেখাবে, ব্রাঞ্চ ম্যানেজার করলে শুধু তার ব্রাঞ্চ
        if warehouse:
            if user.is_superuser:
                 transactions_list = transactions_list.filter(
                    Q(source_location__warehouse=warehouse) |
                    Q(destination_location__warehouse=warehouse)
                )

    # --- পেজিনেশন লজিক ---
    paginator = Paginator(transactions_list, 20)  # প্রতি পৃষ্ঠায় ২০টি লেনদেন
    page_number = request.GET.get('page')
    transactions = paginator.get_page(page_number)

    # ব্রাঞ্চ ম্যানেজারের জন্য ফিল্টার ফর্মের ওয়্যারহাউস ফিল্ড محدود করা
    if not user.is_superuser:
        user_warehouse = getattr(user, 'warehouse', None)
        if user_warehouse:
            form.fields['warehouse'].queryset = Warehouse.objects.filter(pk=user_warehouse.pk)
            form.fields['warehouse'].initial = user_warehouse
            form.fields['warehouse'].widget.attrs['disabled'] = True


    context = {
        'title': 'Stock Movement Report',
        'transactions': transactions,
        'form': form,
    }
    return render(request, 'stock/stock_movement_report.html', context)

@login_required
def get_available_lots(request):
    product_id = request.GET.get('product_id')
    location_id = request.GET.get('location_id')
    
    if not product_id or not location_id:
        return JsonResponse([], safe=False)

    lots = LotSerialNumber.objects.filter(
        product_id=product_id, 
        location_id=location_id,
        quantity__gt=0
    ).values('id', 'lot_number', 'quantity')
    
    return JsonResponse(list(lots), safe=False)

def get_lots_by_location_and_product(request):
    product_id = request.GET.get('product_id')
    location_id = request.GET.get('location_id')
    
    lots = LotSerialNumber.objects.filter(
        product_id=product_id,
        location_id=location_id,
        quantity__gt=0
    ).values('id', 'lot_number', 'quantity')
    
    return JsonResponse(list(lots), safe=False)

@login_required
def download_transaction_report(request):
    # transaction_list ভিউ থেকে ফিল্টারিং লজিকটি এখানেও ব্যবহার করা হয়েছে
    transactions_queryset = InventoryTransaction.objects.select_related(
        'product', 'user', 'source_location__warehouse', 'destination_location__warehouse'
    ).order_by('-transaction_date')

    user = request.user
    if not user.is_superuser:
        user_warehouse = getattr(user, 'warehouse', None)
        if user_warehouse:
            transactions_queryset = transactions_queryset.filter(
                Q(source_location__warehouse=user_warehouse) | Q(destination_location__warehouse=user_warehouse)
            )

    # GET প্যারামিটার থেকে ফিল্টার করা (ফর্ম ব্যবহার না করে সরাসরি)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    selected_user_id = request.GET.get('user')
    warehouse_id = request.GET.get('warehouse')

    if start_date:
        transactions_queryset = transactions_queryset.filter(transaction_date__date__gte=start_date)
    if end_date:
        transactions_queryset = transactions_queryset.filter(transaction_date__date__lte=end_date)
    if selected_user_id:
        transactions_queryset = transactions_queryset.filter(user_id=selected_user_id)
    if warehouse_id and user.is_superuser:
        transactions_queryset = transactions_queryset.filter(
            Q(source_location__warehouse_id=warehouse_id) | Q(destination_location__warehouse_id=warehouse_id)
        )
    
    # --- Excel ফাইল তৈরি ---
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="transaction_report.xlsx"'
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Transactions"

    headers = ['Date', 'Product', 'Type', 'Quantity', 'User', 'Source', 'Destination', 'Notes']
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for transaction in transactions_queryset:
        worksheet.append([
            transaction.transaction_date.strftime('%Y-%m-%d %H:%M'),
            transaction.product.name,
            transaction.get_transaction_type_display(),
            transaction.quantity,
            transaction.user.username if transaction.user else 'N/A',
            f"{transaction.source_location.name} ({transaction.source_location.warehouse.name})" if transaction.source_location else 'N/A',
            f"{transaction.destination_location.name} ({transaction.destination_location.warehouse.name})" if transaction.destination_location else 'N/A',
            transaction.notes or ''
        ])

    for col_num, column_title in enumerate(headers, 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = 20

    workbook.save(response)
    return response

@login_required
def product_stock_details(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    user = request.user
    
    # --- ধাপ ক: সব ব্রাঞ্চের স্টকের সারসংক্ষেপ (সবার জন্য) ---
    # ================== পরিবর্তিত অংশ ==================
    # এখানে 'location__warehouse__name' এর পরিবর্তে সরাসরি 'warehouse__name' ব্যবহার করা হয়েছে
    stock_by_warehouse = Stock.objects.filter(product=product, quantity__gt=0).values(
        'warehouse__name'
    ).annotate(
        total_stock=Sum('quantity')
    ).order_by('warehouse__name')
    # =================================================

    # --- ধাপ খ: ব্যবহারকারীর ভূমিকা অনুযায়ী লটের বিস্তারিত তথ্য (অপরিবর্তিত) ---
    lot_details = None
    user_warehouse = getattr(user, 'warehouse', None)
    
    if user.is_superuser:
        # সুপারইউজার এই প্রোডাক্টের সব ব্রাঞ্চের লট দেখতে পাবে
        lot_details = LotSerialNumber.objects.filter(
            product=product, 
            quantity__gt=0
        ).select_related('location', 'location__warehouse').order_by('location__warehouse__name', 'expiration_date')
    
    # শুধুমাত্র 'Branch Manager' গ্রুপে থাকলেই এই অংশটি কাজ করবে
    elif user.groups.filter(name='Branch Manager').exists() and user_warehouse:
        # ব্রাঞ্চ ম্যানেজার শুধুমাত্র তার নিজের ব্রাঞ্চের লট দেখতে পাবে
        lot_details = LotSerialNumber.objects.filter(
            product=product, 
            location__warehouse=user_warehouse,
            quantity__gt=0
        ).select_related('location').order_by('expiration_date')

    # আপনার context ডিকশনারি অপরিবর্তিত রাখা হয়েছে
    context = {
        'title': f'Stock Details for {product.name}',
        'product': product,
        'stock_by_warehouse': stock_by_warehouse,
        'lot_details': lot_details,
        'user_warehouse': user_warehouse
    }
    return render(request, 'stock/product_stock_details.html', context)