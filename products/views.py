# products/views.py

import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.db.models import Sum, Count, F, Q
from django.db.models.functions import Coalesce
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from datetime import timedelta
from django.db.models.functions import TruncMonth # For monthly trend charts
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import IntegrityError

# Local Application Imports
from .models import Product, Category, UnitOfMeasureCategory, UnitOfMeasure, Brand
from .forms import ProductForm, CategoryForm, UnitOfMeasureCategoryForm, UnitOfMeasureForm, BrandForm
from sales.models import SalesOrder
from partners.models import Supplier, Customer 
from products.models import Product
from stock.models import Stock
from stock.models import Warehouse

# Standard Library Imports
from io import BytesIO
import os

from .admin import ProductResource
from .forms import ProductImportForm
from tablib import Dataset

# Third-Party Imports for Excel Export
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Alignment
from PIL import Image as PillowImage

# Third-Party Imports for PDF Export
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

DEFAULT_CURRENCY_SYMBOL = 'QAR '

def apply_product_filters(request, base_queryset):
    """
    এই কেন্দ্রীয় ফাংশনটি এখন স্টক স্ট্যাটাস অনুযায়ী সঠিকভাবে ফিল্টার করবে।
    """
    query = request.GET.get('q')
    category_id = request.GET.get('category')
    brand_id = request.GET.get('brand')
    status = request.GET.get('status')
    warehouse_id = request.GET.get('warehouse')

    filtered_queryset = base_queryset
    
    if query:
        filtered_queryset = filtered_queryset.filter(Q(name__icontains=query) | Q(product_code__icontains=query))
    if category_id:
        filtered_queryset = filtered_queryset.filter(category_id=category_id)
    if brand_id:
        filtered_queryset = filtered_queryset.filter(brand_id=brand_id)
    if warehouse_id:
        filtered_queryset = filtered_queryset.filter(stocks__warehouse_id=warehouse_id)

    # --- নতুন এবং উন্নত স্ট্যাটাস ফিল্টারিং লজিক ---
    if status:
        if status == 'in_stock':
            filtered_queryset = filtered_queryset.filter(calculated_total_quantity__gt=F('min_stock_level'))
        elif status == 'low_stock':
            filtered_queryset = filtered_queryset.filter(
                calculated_total_quantity__lte=F('min_stock_level'),
                calculated_total_quantity__gt=0
            )
        elif status == 'out_of_stock':
            filtered_queryset = filtered_queryset.filter(calculated_total_quantity__lte=0)
    return filtered_queryset.distinct()

# ধাপ ২: এই ফাংশনটিকেও নিচের নতুন কোড দিয়ে প্রতিস্থাপন করুন
@login_required
@permission_required('products.view_product', login_url='/admin/')
def product_list(request):
    user = request.user
    user_warehouse = getattr(user, 'warehouse', None)

    # বেস কোয়েরিসেট, যেখানে quantity গণনা করা হয়েছে
    products_query = Product.objects.select_related('category', 'brand')
    
    quantity_annotation_filter = Q()
    if not user.is_superuser and user_warehouse:
        quantity_annotation_filter = Q(stocks__warehouse=user_warehouse)
    
    products_query = products_query.annotate(
        calculated_total_quantity=Coalesce(Sum('stocks__quantity', filter=quantity_annotation_filter), 0)
    ).order_by('name')

    # --- নতুন: কেন্দ্রীয় ফিল্টার ফাংশনকে কল করা হয়েছে ---
    filtered_products = apply_product_filters(request, products_query)

    # পেজিনেশন
    paginator = Paginator(filtered_products, 15) # আপনার পছন্দমত সংখ্যা দিন
    page_number = request.GET.get('page')
    products_page = paginator.get_page(page_number)
    
    context = {
        'products': products_page,
        'title': 'All Products',
        'all_categories': Category.objects.all().order_by('name'),
        'all_brands': Brand.objects.all().order_by('name'),
        'product_statuses': Product.STATUS_CHOICES,
        'all_warehouses': Warehouse.objects.all().order_by('name'),
        'DEFAULT_CURRENCY_SYMBOL': 'QAR ', # আপনার কারেন্সি সিম্বল
    }
    return render(request, 'products/product_list.html', context)

@login_required
@permission_required('products.add_product', login_url='/admin/')
def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('products:product_list')
    else:
        form = ProductForm()
    return render(request, 'products/add_product.html', {'form': form, 'title': 'Add New Product'})

@login_required
@permission_required('products.change_product', login_url='/admin/')
def edit_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            return redirect('products:product_list')
    else:
        form = ProductForm(instance=product)
    return render(request, 'products/edit_product.html', {'form': form, 'title': f'Edit Product: {product.name}'})

@login_required
@permission_required('products.delete_product', login_url='/admin/')
def delete_product(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        return redirect('products:product_list')
    return render(request, 'confirm_delete.html', {'object': product, 'title': f'Confirm Delete: {product.name}'})

# --- Category Views ---
@login_required
@permission_required('products.view_category', login_url='/admin/')
def category_list(request):
    categories = Category.objects.all().order_by('name')
    return render(request, 'products/category_list.html', {'categories': categories, 'title': 'All Categories'})

@login_required
@permission_required('products.add_category', login_url='/admin/')
def add_category(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('products:category_list')
    else:
        form = CategoryForm()
    return render(request, 'products/add_category.html', {'form': form, 'title': 'Add New Category'})

@login_required
@permission_required('products.change_category', login_url='/admin/')
def edit_category(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect('products:category_list')
    else:
        form = CategoryForm(instance=category)
    return render(request, 'products/edit_category.html', {'form': form, 'title': f'Edit Category: {category.name}'})

@login_required
@permission_required('products.delete_category', login_url='/admin/')
def delete_category(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        category.delete()
        return redirect('products:category_list')
    return render(request, 'confirm_delete.html', {'object': category, 'title': 'Confirm Delete Category'})

# --- Unit of Measure (UoM) Category Views ---
@login_required
@permission_required('products.view_unitofmeasurecategory', login_url='/admin/')
def uom_category_list(request):
    uom_categories = UnitOfMeasureCategory.objects.all().order_by('name')
    return render(request, 'products/uom_category_list.html', {'uom_categories': uom_categories, 'title': 'Unit of Measure Categories'})

@login_required
@permission_required('products.add_unitofmeasurecategory', login_url='/admin/')
def add_uom_category(request):
    if request.method == 'POST':
        form = UnitOfMeasureCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('products:uom_category_list')
    else:
        form = UnitOfMeasureCategoryForm()
    return render(request, 'products/add_uom_category.html', {'form': form, 'title': 'Add New UoM Category'})

@login_required
@permission_required('products.change_unitofmeasurecategory', login_url='/admin/')
def edit_uom_category(request, pk):
    uom_category = get_object_or_404(UnitOfMeasureCategory, pk=pk)
    if request.method == 'POST':
        form = UnitOfMeasureCategoryForm(request.POST, instance=uom_category)
        if form.is_valid():
            form.save()
            return redirect('products:uom_category_list')
    else:
        form = UnitOfMeasureCategoryForm(instance=uom_category)
    return render(request, 'products/edit_uom_category.html', {'form': form, 'title': f'Edit UoM Category: {uom_category.name}'})

@login_required
@permission_required('products.delete_unitofmeasurecategory', login_url='/admin/')
def delete_uom_category(request, pk):
    uom_category = get_object_or_404(UnitOfMeasureCategory, pk=pk)
    if request.method == 'POST':
        uom_category.delete()
        return redirect('products:uom_category_list')
    return render(request, 'confirm_delete.html', {'object': uom_category, 'title': 'Confirm Delete UoM Category'})

# --- Unit of Measure (UoM) Views ---
@login_required
@permission_required('products.view_unitofmeasure', login_url='/admin/')
def unit_of_measure_list(request):
    units_of_measure = UnitOfMeasure.objects.all().order_by('name')
    context = {'units_of_measure': units_of_measure, 'title': 'Units of Measure'}
    return render(request, 'products/unit_of_measure_list.html', context)

@login_required
@permission_required('products.add_unitofmeasure', login_url='/admin/')
def add_unit_of_measure(request):
    if request.method == 'POST':
        form = UnitOfMeasureForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('products:unit_of_measure_list')
    else:
        form = UnitOfMeasureForm()
    return render(request, 'products/add_unit_of_measure.html', {'form': form, 'title': 'Add New Unit of Measure'})

@login_required
@permission_required('products.change_unitofmeasure', login_url='/admin/')
def edit_unit_of_measure(request, pk):
    unit_of_measure = get_object_or_404(UnitOfMeasure, pk=pk)
    if request.method == 'POST':
        form = UnitOfMeasureForm(request.POST, instance=unit_of_measure)
        if form.is_valid():
            form.save()
            return redirect('products:unit_of_measure_list')
    else:
        form = UnitOfMeasureForm(instance=unit_of_measure)
    return render(request, 'products/edit_unit_of_measure.html', {'form': form, 'title': f'Edit Unit of Measure: {unit_of_measure.name}'})

@login_required
@permission_required('products.delete_unitofmeasure', login_url='/admin/')
def delete_unit_of_measure(request, pk):
    unit_of_measure = get_object_or_404(UnitOfMeasure, pk=pk)
    if request.method == 'POST':
        unit_of_measure.delete()
        return redirect('products:unit_of_measure_list')
    return render(request, 'confirm_delete.html', {'object': unit_of_measure, 'title': 'Confirm Delete Unit of Measure'})

# --- AJAX Views ---
def get_product_sale_price(request):
    product_id = request.GET.get('product_id')
    product = get_object_or_404(Product, id=product_id)
    return JsonResponse({'sale_price': product.sale_price})

@login_required
@permission_required('products.view_product', login_url='/admin/')
def export_products_excel(request):
    # --- নতুন কোড শুরু ---
    user = request.user
    user_warehouse = getattr(user, 'warehouse', None)

    # ব্রাঞ্চ অনুযায়ী স্টক গণনার জন্য ফিল্টার তৈরি
    quantity_annotation_filter = Q()
    if not user.is_superuser and user_warehouse:
        quantity_annotation_filter = Q(stocks__warehouse=user_warehouse)
    # --- নতুন কোড শেষ ---

    products_query = Product.objects.select_related('category', 'brand').annotate(
        # --- পরিবর্তিত লাইন ---
        # এখানে ফিল্টারটি প্রয়োগ করা হয়েছে
        calculated_total_quantity=Coalesce(Sum('stocks__quantity', filter=quantity_annotation_filter), 0)
    ).order_by('name')

    # নিচের অংশ প্রায় অপরিবর্তিত থাকবে
    filtered_products = apply_product_filters(request, products_query)

    report_branch_info = "All Branches"
    warehouse_id = request.GET.get('warehouse')

    if user.is_superuser and warehouse_id:
        try:
            selected_warehouse = Warehouse.objects.get(id=warehouse_id)
            report_branch_info = f"Branch: {selected_warehouse.name}"
        except Warehouse.DoesNotExist:
            pass 
    elif not user.is_superuser and user_warehouse:
        report_branch_info = f"Branch: {user_warehouse.name}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Products Inventory"

    ws.merge_cells('A1:I1')
    report_title_cell = ws['A1']
    report_title_cell.value = f"Product Inventory Report ({report_branch_info})"
    report_title_cell.font = Font(bold=True, size=16)
    report_title_cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:I2')
    import datetime
    ws['A2'].value = f"Report generated on: {datetime.date.today().strftime('%d-%b-%Y')}"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.append([])

    headers = ['Product Name', 'SKU', 'Category', 'Brand', 'Purchase Price', 'Cost Price', 'Sale Price', 'Stock Qty', 'Status']
    ws.append(headers)
    for cell in ws[4]:
        cell.font = Font(bold=True, size=12)
        from openpyxl.styles import PatternFill
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    for product in filtered_products:
        ws.append([
            product.name,
            product.product_code,
            product.category.name if product.category else 'N/A',
            product.brand.name if product.brand else 'N/A',
            product.price,
            product.cost_price,
            product.sale_price,
            product.calculated_total_quantity,
            product.get_status_display(),
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="products_inventory.xlsx"'
    wb.save(response)
    
    return response


@login_required
@permission_required('products.view_product', login_url='/admin/')
def export_products_pdf(request):
    user = request.user
    user_warehouse = getattr(user, 'warehouse', None)

    quantity_annotation_filter = Q()
    if not user.is_superuser and user_warehouse:
        quantity_annotation_filter = Q(stocks__warehouse=user_warehouse)

    products_query = Product.objects.select_related('category', 'brand').annotate(
        calculated_total_quantity=Coalesce(Sum('stocks__quantity', filter=quantity_annotation_filter), 0)
    ).order_by('name')

    filtered_products = apply_product_filters(request, products_query)
    
    report_branch_info = "All Branches"
    warehouse_id = request.GET.get('warehouse')

    if user.is_superuser and warehouse_id:
        try:
            selected_warehouse = Warehouse.objects.get(id=warehouse_id)
            report_branch_info = f"Branch: {selected_warehouse.name}"
        except Warehouse.DoesNotExist:
            pass 
    elif not user.is_superuser and user_warehouse:
        report_branch_info = f"Branch: {user_warehouse.name}"

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=50, bottomMargin=50)
    styles = getSampleStyleSheet()
    story = []

    styles.add(ParagraphStyle(name='ReportHeading', fontSize=18, alignment=TA_CENTER, fontName='Helvetica-Bold', spaceAfter=5))
    styles.add(ParagraphStyle(name='SubHeading', fontSize=10, alignment=TA_CENTER, spaceAfter=20))
    styles.add(ParagraphStyle(name='TableHeader', fontSize=8, alignment=TA_CENTER, fontName='Helvetica-Bold'))
    styles.add(ParagraphStyle(name='TableCell', fontSize=7, alignment=TA_CENTER))
    styles.add(ParagraphStyle(name='TableCellLeft', fontSize=7, alignment=TA_LEFT))

    story.append(Paragraph("Product Inventory Report", styles['ReportHeading']))
    story.append(Paragraph(f"({report_branch_info})", styles['SubHeading']))
    
    table_data = [
        [Paragraph(h, styles['TableHeader']) for h in ['Product Name', 'SKU', 'Category', 'Brand', 'Cost Price', 'Sale Price', 'Qty']]
    ]
    
    for product in filtered_products:
        table_data.append([
            Paragraph(product.name, styles['TableCellLeft']),
            Paragraph(product.product_code or 'N/A', styles['TableCell']),
            Paragraph(product.category.name if product.category else 'N/A', styles['TableCell']),
            Paragraph(product.brand.name if product.brand else 'N/A', styles['TableCell']),
            Paragraph(f"{product.cost_price:.2f}", styles['TableCell']),
            Paragraph(f"{product.sale_price:.2f}", styles['TableCell']),
            Paragraph(str(product.calculated_total_quantity), styles['TableCell']),
        ])

    table = Table(table_data, colWidths=[2.5*inch, 1*inch, 1*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        # --- এই লাইনটিতে বানান ঠিক করা হয়েছে ---
        ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke), 
        # --- whitesmokoke এর পরিবর্তে whitesmoke ---
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    
    story.append(table)
    doc.build(story)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="products_inventory.pdf"'
    return response

@login_required
def print_product_labels(request):
    if request.method == 'POST':
        product_ids = request.POST.getlist('product_ids')

        if not product_ids:
            return redirect('products:product_list')

        labels_to_print = []
        for pid in product_ids:
            try:
                product = Product.objects.get(id=int(pid))
                quantity = int(request.POST.get(f'quantity_{pid}', 1))
                
                if product.barcode and hasattr(product.barcode, 'path'):
                    for _ in range(quantity):
                        labels_to_print.append({
                            'name': product.name,
                            'price': f"{product.sale_price:.2f}",
                            'currency': DEFAULT_CURRENCY_SYMBOL,
                            'barcode_path': product.barcode.path,
                            'product_code': product.product_code
                        })
            except (Product.DoesNotExist, ValueError):
                continue

        if not labels_to_print:
            return redirect('products:product_list')

        # --- Thermal Printer-এর জন্য PDF তৈরির প্রক্রিয়া ---
        buffer = BytesIO()
        label_width, label_height = 2.25 * inch, 1.25 * inch
        doc = SimpleDocTemplate(buffer, pagesize=(label_width, label_height), leftMargin=0.05*inch, rightMargin=0.05*inch, topMargin=0.05*inch, bottomMargin=0.05*inch)
        story = []
        
        for label_data in labels_to_print:
            product_name = label_data['name']
            if len(product_name) > 30:
                 name_style = ParagraphStyle('LabelName', fontSize=6, alignment=TA_CENTER, leading=7)
            else:
                 name_style = ParagraphStyle('LabelName', fontSize=7, alignment=TA_CENTER, leading=8)

            inner_table_data = [
                [Paragraph(product_name, name_style)],
                [RLImage(label_data['barcode_path'], width=1.8*inch, height=0.4*inch)],
                [
                    Paragraph(label_data['product_code'], ParagraphStyle('LabelCode', fontName='Helvetica', fontSize=7, alignment=TA_LEFT)),
                    Paragraph(f"<b>{label_data['currency']}{label_data['price']}</b>", ParagraphStyle('LabelPrice', fontName='Helvetica-Bold', fontSize=10, alignment=TA_RIGHT))
                ]
            ]
            
            inner_table = Table(inner_table_data, 
                                colWidths=[1.025*inch, 1.025*inch], 
                                rowHeights=[0.25*inch, 0.45*inch, 0.2*inch])
            
            inner_table.setStyle(TableStyle([
                ('SPAN', (0,0), (1,0)), ('SPAN', (0,1), (1,1)),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (0,2), (0,2), 'LEFT'), ('ALIGN', (1,2), (1,2), 'RIGHT'),
                ('VALIGN', (0,2), (1,2), 'BOTTOM'),
            ]))
            
            story.append(inner_table)

        doc.build(story)
        buffer.seek(0)
        return HttpResponse(buffer, content_type='application/pdf')

    return redirect('products:product_list')

@login_required
@permission_required('products.change_unitofmeasurecategory', login_url='/admin/')
def edit_uom_category(request, pk):
    uom_category = get_object_or_404(UnitOfMeasureCategory, pk=pk)
    if request.method == 'POST':
        form = UnitOfMeasureCategoryForm(request.POST, instance=uom_category)
        if form.is_valid():
            form.save()
            return redirect('products:uom_category_list')
    else:
        form = UnitOfMeasureCategoryForm(instance=uom_category)
    return render(request, 'products/edit_uom_category.html', {'form': form, 'title': f'Edit UoM Category: {uom_category.name}'})

@login_required
@permission_required('products.delete_unitofmeasurecategory', login_url='/admin/')
def delete_uom_category(request, pk):
    uom_category = get_object_or_404(UnitOfMeasureCategory, pk=pk)
    if request.method == 'POST':
        uom_category.delete()
        return redirect('products:uom_category_list')
    return render(request, 'confirm_delete.html', {'object': uom_category, 'title': 'Confirm Delete UoM Category'})

@login_required
@permission_required('stock.view_stock', login_url='/admin/')
def product_stock_by_location(request, pk):
    product = get_object_or_404(Product, pk=pk)
    stocks = Stock.objects.filter(product=product).select_related('location').order_by('location__name')
    
    total_stock = sum(s.quantity for s in stocks)
    
    context = {
        'title': f'Stock for {product.name}',
        'product': product,
        'stocks': stocks,
        'total_stock': total_stock,
    }
    return render(request, 'products/product_stock_by_location.html', context)

@login_required
@permission_required('products.view_brand', login_url='/admin/')
def brand_list(request):
    brands = Brand.objects.all().order_by('name')
    return render(request, 'products/brand_list.html', {'brands': brands, 'title': 'All Brands'})

@login_required
@permission_required('products.add_brand', login_url='/admin/')
def add_brand(request):
    if request.method == 'POST':
        form = BrandForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Brand added successfully!')
            return redirect('products:brand_list')
    else:
        form = BrandForm()
    return render(request, 'products/add_brand.html', {'form': form, 'title': 'Add New Brand'})

@login_required
@permission_required('products.change_brand', login_url='/admin/')
def edit_brand(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    if request.method == 'POST':
        form = BrandForm(request.POST, instance=brand)
        if form.is_valid():
            form.save()
            messages.success(request, f'Brand "{brand.name}" updated successfully!')
            return redirect('products:brand_list')
    else:
        form = BrandForm(instance=brand)
    return render(request, 'products/edit_brand.html', {'form': form, 'title': f'Edit Brand: {brand.name}'})

@login_required
@permission_required('products.delete_brand', login_url='/admin/')
def delete_brand(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    if request.method == 'POST':
        brand.delete()
        messages.success(request, f'Brand "{brand.name}" deleted successfully!')
        return redirect('products:brand_list')
    return render(request, 'confirm_delete.html', {'object': brand, 'title': 'Confirm Delete Brand'})

@login_required
@permission_required('products.change_product', login_url='/admin/')
def product_bulk_action(request):
    if request.method == 'POST':
        product_ids = request.POST.getlist('product_ids')
        action = request.POST.get('action')

        if not product_ids:
            messages.error(request, "Please select at least one product.")
            return redirect('products:product_list')

        if action == 'activate':
            Product.objects.filter(id__in=product_ids).update(is_active=True)
            messages.success(request, f"{len(product_ids)} product(s) have been activated.")
        elif action == 'deactivate':
            Product.objects.filter(id__in=product_ids).update(is_active=False)
            messages.success(request, f"{len(product_ids)} product(s) have been deactivated.")
        else:
            messages.warning(request, "Invalid action selected.")

    return redirect('products:product_list')

@login_required
@permission_required('products.add_product', login_url='/admin/')
def import_products_view(request):
    if request.method == 'POST':
        form = ProductImportForm(request.POST, request.FILES)
        if form.is_valid():
            product_resource = ProductResource()
            dataset = Dataset()
            new_products_file = request.FILES['file']

            try:
                if new_products_file.name.endswith('csv'):
                    dataset.load(new_products_file.read().decode('utf-8'), format='csv')
                elif new_products_file.name.endswith(('xls', 'xlsx')):
                    dataset.load(new_products_file.read(), format='xlsx')
                else:
                    messages.error(request, "Unsupported file format. Please upload a CSV or Excel file.")
                    return redirect('products:import_products')

                # 'dry_run' ব্যবহার করে প্রথমে ডেটা যাচাই করা হচ্ছে
                result = product_resource.import_data(dataset, dry_run=True, raise_errors=True)

                # কোনো ভুল না থাকলে ডেটা সেভ করা হচ্ছে
                product_resource.import_data(dataset, dry_run=False)
                messages.success(request, "Products imported successfully!")
                return redirect('products:product_list')

            except IntegrityError as e:
                messages.error(request, f"Data integrity error: {e}")
            except Exception as e:
                # ফাইলের ভেতরের ভুলের জন্য
                error_html = "<ul>"
                for error in getattr(e, 'errors', []):
                     error_html += f"<li>Row {error.row}: {', '.join(error.error.messages)}</li>"
                if not getattr(e, 'errors', []):
                    error_html += f"<li>An unexpected error occurred: {e}</li>"
                error_html += "</ul>"
                messages.error(request, f"There were errors in the file:{error_html}")


    else:
        form = ProductImportForm()

    context = {
        'form': form,
        'title': 'Import Products'
    }
    return render(request, 'products/import_products.html', context)

@login_required
def download_sample_template(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Product Import Template"

    # কলাম হেডারগুলো যোগ করা হচ্ছে
    headers = [
        'id', 'name', 'product_code', 'category__name', 'brand__name', 'price', 
        'sale_price', 'cost_price', 'min_stock_level', 'tracking_method', 'is_active'
    ]
    sheet.append(headers)
    
    # হেডার সেলগুলোকে বোল্ড করা হচ্ছে
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    # --- নতুন: নমুনা ডেটা যোগ করা হলো ---
    sample_data = [
        ['', 'New Laptop Model', 'LP-105', 'Electronics', 'HP', 55000, 65000, 52000, 10, 'serial', 1],
        ['', 'Office Chair', 'CHR-02', 'Furniture', 'Regal', 4500, 6000, 4000, 15, 'none', 1],
        [12, 'Existing Product Name', 'EXT-001', 'Existing Category', 'Existing Brand', 250, 300, 220, 5, 'none', 0],
        ['', 'Product Without Brand', 'PWB-01', 'General', '', 100, 120, 90, 25, 'none', 1],
    ]

    for row_data in sample_data:
        sheet.append(row_data)

    # কলামের প্রস্থ অ্যাডজাস্ট করা হচ্ছে
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

    # HTTP response তৈরি করা হচ্ছে
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="product_import_template.xlsx"'
    
    workbook.save(response)
    
    return response