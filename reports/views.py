# reports/views.py (সংশোধিত)

import os
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q
from django.conf import settings
from datetime import timedelta, datetime
from django.http import HttpResponse
from django.contrib.auth import get_user_model
from django.core.paginator import Paginator
from django.utils import timezone
from django.utils.dateparse import parse_date
from io import BytesIO

# মডেল ইম্পোর্ট
from sales.models import SalesOrder, SalesReturn
from stock.models import Warehouse, LotSerialNumber, Stock
from sales.models import SalesOrderItem
from products.models import Product

# এক্সেল এবং পিডিএফ তৈরির লাইব্রেরি
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.units import points_to_pixels
import openpyxl
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT


User = get_user_model()
DEFAULT_CURRENCY_SYMBOL = 'QAR '

def get_daily_ledger_data(request):
    """
    এই কেন্দ্রীয় ফাংশনটি request থেকে ফিল্টার নেয় এবং সেলস ও রিটার্নের সমন্বিত তালিকা 
    এবং মোট হিসাব প্রদান করে। ওয়েবপেজ ও ডাউনলোড ফাংশন এখন এটি ব্যবহার করবে।
    """
    user = request.user

    # --- ১. ফর্ম থেকে ফিল্টারের তথ্য গ্রহণ ---
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    user_id = request.GET.get('user')
    warehouse_id = request.GET.get('warehouse')

    # --- নতুন: ডিফল্ট তারিখের যুক্তি ---
    form_start_date = parse_date(start_date_str) if start_date_str else None
    form_end_date = parse_date(end_date_str) if end_date_str else None

    query_start_date, query_end_date = form_start_date, form_end_date

    # যদি কোনো তারিখ সিলেক্ট করা না থাকে, তবে আজকের তারিখ ডিফল্ট হিসেবে ধরবে
    if not form_start_date and not form_end_date:
        today = timezone.now().date()
        query_start_date, query_end_date = today, today

    # --- ২. কুয়েরিসেট তৈরি ও ফিল্টার ---
    sales_qs = SalesOrder.objects.filter(status='delivered').select_related('customer', 'user', 'warehouse')
    returns_qs = SalesReturn.objects.select_related('sales_order__customer', 'user', 'warehouse').prefetch_related('items')

    # ব্যবহারকারীর ভূমিকা অনুযায়ী ওয়্যারহাউস ফিল্টার
    if not user.is_superuser:
        user_warehouse = getattr(user, 'warehouse', None)
        if user_warehouse:
            sales_qs = sales_qs.filter(warehouse=user_warehouse)
            returns_qs = returns_qs.filter(warehouse=user_warehouse)

    # ফর্ম থেকে আসা ফিল্টার প্রয়োগ
    if query_start_date:
        sales_qs = sales_qs.filter(order_date__date__gte=query_start_date)
        returns_qs = returns_qs.filter(return_date__date__gte=query_start_date)
    if query_end_date:
        sales_qs = sales_qs.filter(order_date__date__lte=query_end_date)
        returns_qs = returns_qs.filter(return_date__date__lte=query_end_date)
    if user_id:
        sales_qs = sales_qs.filter(user_id=user_id)
        returns_qs = returns_qs.filter(user_id=user_id)
    if user.is_superuser and warehouse_id:
        sales_qs = sales_qs.filter(warehouse_id=warehouse_id)
        returns_qs = returns_qs.filter(warehouse_id=warehouse_id)

    # --- ৩. সেলস এবং রিটার্নকে একটি লিস্টে একত্রিত করা ---
    daily_ledger = []
    for sale in sales_qs:
        daily_ledger.append({
            'date': sale.order_date,
            'type': 'Sale',
            'order_obj': sale,
            'return_obj': None,
            'amount': sale.total_amount
        })
    
    for ret in returns_qs:
        return_amount = sum(item.subtotal for item in ret.items.all())
        daily_ledger.append({
            'date': ret.return_date,
            'type': 'Return',
            'order_obj': ret.sales_order,
            'return_obj': ret,
            'amount': -return_amount
        })
        
    daily_ledger.sort(key=lambda x: x['date'], reverse=True)

    # --- ৪. মোট হিসাব গণনা ---
    total_sales = sales_qs.aggregate(total=Sum('total_amount'))['total'] or 0
    total_returns = sum(sum(item.subtotal for item in r.items.all()) for r in returns_qs)
    net_sales = total_sales - total_returns

    return {
        'daily_ledger': daily_ledger,
        'total_sales': total_sales,
        'total_returns': total_returns,
        'net_sales': net_sales,
        'form_start_date': form_start_date, # টেমপ্লেটের জন্য
        'form_end_date': form_end_date,     # টেমপ্লেটের জন্য
    }

def daily_sales_report(request):
    user = request.user
    report_data = get_daily_ledger_data(request)

    paginator = Paginator(report_data['daily_ledger'], 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    users_qs = User.objects.all()
    warehouses_qs = Warehouse.objects.all()
    if not user.is_superuser:
        user_warehouse = getattr(user, 'warehouse', None)
        if user_warehouse:
            warehouses_qs = Warehouse.objects.filter(pk=user_warehouse.pk)
            users_qs = User.objects.filter(warehouse=user_warehouse)

    context = {
        'title': 'Daily Sales Report',
        'daily_ledger': page_obj,
        'total_sales': report_data['total_sales'],
        'total_returns': report_data['total_returns'],
        'net_sales': report_data['net_sales'],
        'form_start_date': report_data['form_start_date'], # context এ যোগ করা হয়েছে
        'form_end_date': report_data['form_end_date'],     # context এ যোগ করা হয়েছে
        'users': users_qs,
        'warehouses': warehouses_qs,
        'DEFAULT_CURRENCY_SYMBOL': DEFAULT_CURRENCY_SYMBOL,
    }
    return render(request, 'reports/daily_sales_report.html', context)


def export_daily_sales_excel(request):
    report_data = get_daily_ledger_data(request)
    daily_ledger = report_data['daily_ledger']

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Daily_Sales_Report.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Sales Report"

    # --- স্টাইল এবং ফরম্যাটিং ---
    header_font = Font(name='Poppins', size=11, bold=True, color="FFFFFF")
    title_font = Font(name='Poppins', size=18, bold=True, color="2B3674")
    company_font = Font(name='Poppins', size=10, bold=True, color="2B3674")
    total_label_font = Font(name='Poppins', size=10, bold=True)
    
    header_fill = PatternFill(start_color="4E73DF", end_color="4E73DF", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- ১. হেডার সেকশন ---
    ws.merge_cells('A1:B2')
    logo_path = os.path.join(settings.STATICFILES_DIRS[0], 'images', 'logo.png')
    img = openpyxl.drawing.image.Image(logo_path)
    img.height = 40
    img.width = 120
    ws.add_image(img, 'A1')
    
    ws.merge_cells('C1:G1')
    ws['C1'] = "Daily Sales Report"
    ws['C1'].font = title_font
    ws['C1'].alignment = right_align

    ws.merge_cells('C2:G2')
    ws['C2'] = "NOVO ERP Solutions, Doha, Qatar"
    ws['C2'].font = company_font
    ws['C2'].alignment = right_align

    ws.append([]) # একটি খালি সারি

    # --- ২. টেবিল হেডার ---
    headers = ['ID', 'Date', 'Type', 'Customer', 'User', 'Branch', 'Amount']
    ws.append(headers)
    
    header_row = ws[4]
    for cell in header_row:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # --- ৩. ডেটা সারি ---
    for entry in daily_ledger:
        naive_datetime = timezone.localtime(entry['date']).replace(tzinfo=None)
        
        if entry['type'] == 'Sale':
            order = entry['order_obj']
            row_data = [ 
                f"SO-{order.id}", naive_datetime, 'Sale', 
                order.customer.name if order.customer else 'N/A', 
                order.user.username if order.user else 'N/A', 
                order.warehouse.name if order.warehouse else 'N/A', 
                float(entry['amount']) 
            ]
        else: # Return
            ret = entry['return_obj']
            order = entry['order_obj']
            row_data = [ 
                f"RT-{ret.id}", naive_datetime, 'Return', 
                order.customer.name if order.customer else 'N/A', 
                ret.user.username if ret.user else 'N/A', 
                ret.warehouse.name if ret.warehouse else 'N/A', 
                float(entry['amount']) 
            ]
        ws.append(row_data)
        
        # প্রতিটি সেলে বর্ডার এবং ফরম্যাট যোগ করা
        row_index = ws.max_row
        for col_index in range(1, 8):
            cell = ws.cell(row=row_index, column=col_index)
            cell.border = thin_border
            if col_index == 2:
                cell.number_format = 'YYYY-MM-DD HH:MM'
            if col_index == 7:
                cell.number_format = f'"{DEFAULT_CURRENCY_SYMBOL}" #,##0.00'

    # --- ৪. মোট হিসাব সেকশন ---
    last_row = ws.max_row + 2
    ws.cell(row=last_row, column=6, value="Total Sales:").font = total_label_font
    ws.cell(row=last_row, column=6).alignment = right_align
    ws.cell(row=last_row, column=7, value=float(report_data['total_sales'])).number_format = f'"{DEFAULT_CURRENCY_SYMBOL}" #,##0.00'
    
    ws.cell(row=last_row + 1, column=6, value="Total Returns:").font = total_label_font
    ws.cell(row=last_row + 1, column=6).alignment = right_align
    ws.cell(row=last_row + 1, column=7, value=float(report_data['total_returns'])).number_format = f'"{DEFAULT_CURRENCY_SYMBOL}" #,##0.00'
    
    ws.cell(row=last_row + 2, column=6, value="Net Sales:").font = Font(name='Poppins', size=11, bold=True)
    ws.cell(row=last_row + 2, column=6).alignment = right_align
    ws.cell(row=last_row + 2, column=7, value=float(report_data['net_sales'])).number_format = f'"{DEFAULT_CURRENCY_SYMBOL}" #,##0.00'
    ws.cell(row=last_row + 2, column=7, value=float(report_data['net_sales'])).font = Font(name='Poppins', size=11, bold=True)

    # --- ৫. কলামের প্রস্থ ঠিক করা ---
    column_widths = {'A': 12, 'B': 20, 'C': 10, 'D': 30, 'E': 15, 'F': 18, 'G': 18}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(response)
    return response


def export_daily_sales_pdf(request):
    report_data = get_daily_ledger_data(request)
    daily_ledger = report_data['daily_ledger']
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Daily_Sales_Report.pdf"'
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch, bottomMargin=0.5*inch, leftMargin=0.5*inch, rightMargin=0.5*inch)
    story = []

    # --- কাস্টম স্টাইল তৈরি ---
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='TitleStyle', fontSize=22, fontName='Helvetica-Bold', alignment=TA_RIGHT, textColor=colors.HexColor("#2B3674")))
    styles.add(ParagraphStyle(name='CompanyInfo', fontSize=9, fontName='Helvetica', alignment=TA_RIGHT, leading=12))
    styles.add(ParagraphStyle(name='ReportInfo', fontSize=10, fontName='Helvetica', leading=14))
    styles.add(ParagraphStyle(name='TotalLabel', fontSize=10, fontName='Helvetica-Bold', alignment=TA_RIGHT))
    styles.add(ParagraphStyle(name='TotalValue', fontSize=10, fontName='Helvetica-Bold', alignment=TA_RIGHT))
    styles.add(ParagraphStyle(name='SignatureStyle', fontSize=10, fontName='Helvetica', alignment=TA_CENTER))
    
    # --- মূল পরিবর্তন: অনুপস্থিত স্টাইলগুলো এখানে যোগ করা হলো ---
    styles.add(ParagraphStyle(name='TableHeader', fontSize=9, fontName='Helvetica-Bold', alignment=TA_LEFT, textColor=colors.HexColor("#2B3674")))
    styles.add(ParagraphStyle(name='TableCellLeft', fontSize=9, alignment=TA_LEFT))
    styles.add(ParagraphStyle(name='TableCellRight', fontSize=9, alignment=TA_RIGHT))
    styles.add(ParagraphStyle(name='TableCellRightRed', fontSize=9, alignment=TA_RIGHT, textColor=colors.red))


    # --- ১. হেডার সেকশন ---
    logo_path = os.path.join(settings.STATICFILES_DIRS[0], 'images', 'logo.png')
    logo = Image(logo_path, width=1.85*inch, height=0.5*inch)
    company_info = "<b>NOVO ERP Solutions</b><br/>Doha, Qatar"
    
    header_data = [[logo, Paragraph("Daily Sales Report", styles['TitleStyle'])]]
    header_table = Table(header_data, colWidths=[7*inch, 3.5*inch], style=[('VALIGN', (0,0), (-1,-1), 'TOP')])
    story.append(header_table)
    story.append(Spacer(1, 0.05*inch))
    story.append(Paragraph(company_info, styles['CompanyInfo']))
    story.append(Spacer(1, 0.3*inch))

    # --- ২. রিপোর্টের তথ্য ---
    start_date = request.GET.get('start_date', 'N/A')
    end_date = request.GET.get('end_date', 'N/A')
    report_info_text = f"<b>Date Range:</b> {start_date} to {end_date}<br/><b>Report Generated:</b> {timezone.now().strftime('%d %b, %Y %I:%M %p')}"
    story.append(Paragraph(report_info_text, styles['ReportInfo']))
    story.append(Spacer(1, 0.3*inch))

    # --- ৩. টেবিল ---
    table_header = [
        Paragraph('ID', styles['TableHeader']),
        Paragraph('Date', styles['TableHeader']),
        Paragraph('Type', styles['TableHeader']),
        Paragraph('Customer', styles['TableHeader']),
        Paragraph('User', styles['TableHeader']),
        Paragraph('Branch', styles['TableHeader']),
        Paragraph('Amount', styles['TableHeader']),
    ]
    table_data = [table_header]
    
    for entry in daily_ledger:
        amount_style = styles['TableCellRightRed'] if entry['type'] == 'Return' else styles['TableCellRight']
        
        if entry['type'] == 'Sale':
            order = entry['order_obj']
            row_data = [
                Paragraph(f"SO-{order.id}", styles['TableCellLeft']),
                Paragraph(order.order_date.strftime('%Y-%m-%d %H:%M'), styles['TableCellLeft']),
                Paragraph('Sale', styles['TableCellLeft']),
                Paragraph(order.customer.name if order.customer else 'N/A', styles['TableCellLeft']),
                Paragraph(order.user.username if order.user else 'N/A', styles['TableCellLeft']),
                Paragraph(order.warehouse.name if order.warehouse else 'N/A', styles['TableCellLeft']),
                Paragraph(f"{entry['amount']:,.2f}", amount_style)
            ]
        else: # Return
            ret = entry['return_obj']
            order = entry['order_obj']
            row_data = [
                Paragraph(f"RT-{ret.id}", styles['TableCellLeft']),
                Paragraph(ret.return_date.strftime('%Y-%m-%d %H:%M'), styles['TableCellLeft']),
                Paragraph('Return', styles['TableCellLeft']),
                Paragraph(order.customer.name if order.customer else 'N/A', styles['TableCellLeft']),
                Paragraph(ret.user.username if ret.user else 'N/A', styles['TableCellLeft']),
                Paragraph(ret.warehouse.name if ret.warehouse else 'N/A', styles['TableCellLeft']),
                Paragraph(f"{entry['amount']:,.2f}", amount_style)
            ]
        table_data.append(row_data)
        
    table = Table(table_data, colWidths=[0.8*inch, 1.2*inch, 0.6*inch, 3.2*inch, 1.3*inch, 1.55*inch, 1.85*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#E0E5F2")),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor("#CCCCCC")),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(table)
    story.append(Spacer(1, 0.3*inch))

    # --- ৪. মোট হিসাব সেকশন ---
    total_data = [
        [Paragraph('Total Sales:', styles['TotalLabel']), Paragraph(f"{settings.DEFAULT_CURRENCY_SYMBOL} {report_data['total_sales']:,.2f}", styles['TotalValue'])],
        [Paragraph('Total Returns:', styles['TotalLabel']), Paragraph(f"{settings.DEFAULT_CURRENCY_SYMBOL} {report_data['total_returns']:,.2f}", styles['TotalValue'])],
        [Paragraph('Net Sales:', styles['TotalLabel']), Paragraph(f"{settings.DEFAULT_CURRENCY_SYMBOL} {report_data['net_sales']:,.2f}", styles['TotalValue'])],
    ]
    total_table = Table(total_data, colWidths=[1.5*inch, 2*inch], hAlign='RIGHT')
    story.append(total_table)
    
    # --- ৫. সিগনেচার সেকশন ---
    story.append(Spacer(1, 0.5*inch))
    signature_data = [
        [Paragraph('--------------------------------<br/>Prepared By', styles['SignatureStyle']),
         Paragraph('--------------------------------<br/>Checked By', styles['SignatureStyle']),
         Paragraph('--------------------------------<br/>Approved By', styles['SignatureStyle'])]
    ]
    signature_table = Table(signature_data, colWidths=[3*inch, 3*inch, 3*inch], hAlign='CENTER')
    story.append(signature_table)

    doc.build(story)
    
    buffer.seek(0)
    response.write(buffer.getvalue())
    buffer.close()
    return response

@login_required
def expiry_report_view(request):
    user = request.user
    user_warehouse = getattr(user, 'warehouse', None)
    today = timezone.now().date()
    ninety_days_later = today + timedelta(days=90)

    expiring_lots_qs = LotSerialNumber.objects.filter(
        expiration_date__isnull=False, expiration_date__gte=today,
        expiration_date__lte=ninety_days_later, quantity__gt=0
    ).select_related('product', 'location__warehouse').order_by('expiration_date')

    if not user.is_superuser and user_warehouse:
        expiring_lots_qs = expiring_lots_qs.filter(location__warehouse=user_warehouse)

    expiring_lots_list = list(expiring_lots_qs)
    for lot in expiring_lots_list:
        lot.days_left = (lot.expiration_date - today).days if lot.expiration_date else 'N/A'
            
    paginator = Paginator(expiring_lots_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {'title': 'Products Expiring Soon Report', 'page_obj': page_obj}
    return render(request, 'reports/expiry_report.html', context)

@login_required
def dead_stock_report_view(request):
    user = request.user
    user_warehouse = getattr(user, 'warehouse', None)
    today = timezone.now().date()
    ninety_days_ago = today - timedelta(days=90)
    
    sold_product_ids = SalesOrderItem.objects.filter(sales_order__order_date__gte=ninety_days_ago).values_list('product_id', flat=True).distinct()
    stock_in_lots = LotSerialNumber.objects.filter(quantity__gt=0)
    
    if not user.is_superuser and user_warehouse:
        stock_in_lots = stock_in_lots.filter(location__warehouse=user_warehouse)

    dead_stock_lots = stock_in_lots.exclude(product_id__in=sold_product_ids)
    
    dead_stock_summary = {}
    for lot in dead_stock_lots.select_related('product', 'location__warehouse'):
        key = (lot.product, lot.location.warehouse)
        if key not in dead_stock_summary:
            dead_stock_summary[key] = {'product': lot.product, 'warehouse': lot.location.warehouse, 'quantity': 0}
        dead_stock_summary[key]['quantity'] += lot.quantity
        
    dead_stock_list = sorted(list(dead_stock_summary.values()), key=lambda x: x['product'].name)

    paginator = Paginator(dead_stock_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {'title': 'Dead Stock Report (Not Sold in 90 Days)', 'page_obj': page_obj}
    return render(request, 'reports/dead_stock_report.html', context)

@login_required
def purchase_suggestion_report_view(request):
    user = request.user
    user_warehouse = getattr(user, 'warehouse', None)
    today = timezone.now().date()
    thirty_days_ago = today - timedelta(days=30)
    
    stock_qs = Stock.objects.all()
    if not user.is_superuser and user_warehouse:
        stock_qs = stock_qs.filter(warehouse=user_warehouse)
        
    products_in_stock = stock_qs.filter(quantity__gt=0).values('product_id').annotate(current_stock=Sum('quantity'))
    product_stock_map = {item['product_id']: item['current_stock'] for item in products_in_stock}

    sales_qs = SalesOrderItem.objects.filter(
        sales_order__order_date__gte=thirty_days_ago,
        product_id__in=product_stock_map.keys()
    )
    if not user.is_superuser and user_warehouse:
        sales_qs = sales_qs.filter(sales_order__warehouse=user_warehouse)

    sales_last_30_days = sales_qs.values('product_id').annotate(total_sold=Sum('quantity'))

    suggestion_list = []
    for sale in sales_last_30_days:
        product_id = sale['product_id']
        current_stock = product_stock_map.get(product_id, 0)
        total_sold = sale['total_sold']
        
        if current_stock < total_sold:
            product = Product.objects.get(id=product_id)
            suggestion_list.append({
                'product': product, 'current_stock': current_stock,
                'sold_last_30_days': total_sold, 'suggested_quantity': total_sold - current_stock,
            })
            
    paginator = Paginator(suggestion_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {'title': 'Purchase Suggestion Report', 'page_obj': page_obj}
    return render(request, 'reports/purchase_suggestion_report.html', context)